import re
from typing import Dict, List, Union
import asyncio
import os
from telethon import TelegramClient
from telethon.tl.types import MessageMediaPhoto
import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


class TelegramParser:
    def __init__(self, client: TelegramClient):
        self.client = client

    async def get_messages(
        self, entity: str | int, limit: int = 100
    ) -> List[Dict[str, str]]:
        """Gets messages with media files."""
        messages: List[Dict[str, str]] = []
        messages_data = await self.client.get_messages(entity, limit=limit)
        for message in messages_data:

            first_line = message.text.split("\n")[0].strip()
            message_id = (
                first_line.split()[0]
                if first_line and len(first_line.split()) > 0
                else "No ID"
            )

            messages.append(
                {
                    "telegram_id": message.id,
                    "id": re.sub(r"\*", "", message_id),
                    "message": re.sub(r"\s+", " ", message.message).strip(),
                    "photo": message.media,
                }
            )

        return messages

    async def group_objects(
        self, objects
    ) -> List[Dict[str, Union[str, List[MessageMediaPhoto]]]]:
        grouped = []
        current_group = {"message": "", "photos": []}

        for obj in objects:

            if obj["message"]:
                current_group["message"] = obj
                current_group["photos"].append(obj["photo"])
                grouped.append(current_group)
                current_group = {"message": "", "photos": []}
            else:
                current_group["photos"].append(obj["photo"])

        return grouped

    async def download_photos(
        self, grouped_message: List[Dict[str, Union[str, List[MessageMediaPhoto]]]]
    ):
        os.makedirs("photos", exist_ok=True)
        tasks = []
        for message in grouped_message:
            for photo in message.get("photos"):
                filename = os.path.join(
                    "photos",
                    f"{message['message']['id']}_{photo.photo.id}.jpg",
                )
                task = asyncio.create_task(
                    self.client.download_media(photo, file=filename)
                )
                tasks.append(task)

        return await asyncio.gather(*tasks)


async def create_excel_with_ids_and_text(
    data: List[Dict[str, Union[str, List[MessageMediaPhoto]]]],
    filename: str = "tg_messages.xlsx",
    photos_dir: str = "photos",
) -> None:

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Текст", "Фото"])

    # Собираем все фото и создаем гиперссылки
    if os.path.exists(photos_dir):
        for item in data:
            message_id = item["message"]["id"]
            row = [message_id, item["message"]["message"], ""]  # Поле для гиперссылок
            ws.append(row)

            # Ищем все фото для этого ID
            photo_counter = 1
            for photo_file in os.listdir(photos_dir):
                if photo_file.startswith(f"{message_id}_"):
                    photo_path = os.path.abspath(os.path.join(photos_dir, photo_file))

                    # Создаем кликабельную ссылку
                    link_text = f"Фото {photo_counter}"
                    cell = ws.cell(row=ws.max_row, column=3)
                    cell.value = link_text
                    cell.hyperlink = Hyperlink(
                        ref=cell.coordinate,
                        target=photo_path,
                        tooltip=f"Открыть {photo_file}",
                    )
                    photo_counter += 1

    # Настраиваем ширину колонок
    ws.column_dimensions[get_column_letter(1)].width = 15  # ID
    ws.column_dimensions[get_column_letter(2)].width = 50  # Текст
    ws.column_dimensions[get_column_letter(3)].width = 25  # Фото

    wb.save(filename)
    print(f"✅ Файл {filename} создан с кликабельными ссылками!")


async def main():
    api_id = 23128708
    api_hash = "ee3dfa7067eb520a05bfc749083f9ab0"
    phone_number = "+79130028603"
    chat_entity = "prodajadomov_mirzo_ulugbek"

    async with TelegramClient("s", api_id, api_hash) as client:
        try:
            await client.connect()
            if not await client.is_user_authorized():
                await client.send_code_request(phone_number)
                await client.sign_in(phone_number, input("Enter code: "))

            parser = TelegramParser(client)

            messages = await parser.get_messages(chat_entity, limit=45)
            groups = await parser.group_objects(messages)
            await parser.download_photos(groups)
            await create_excel_with_ids_and_text(groups)
            print(*groups, sep="\n\n")
        except Exception as e:
            logging.error(f"An error occurred: {e}")
        finally:
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
